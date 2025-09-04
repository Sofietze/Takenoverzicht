
import pandas as pd
import streamlit as st
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from fpdf import FPDF
from datetime import datetime

def generate_pdf(titel, df, dfproblem):
    pdf = FPDF()
    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # ➕ Voeg Unicode-lettertype toe
    pdf.add_font("DejaVu", "", "DejaVuSans.ttf", uni=True)
  
    pdf.set_font("DejaVu", "", 12)
    # 📌 Titel
    pdf.cell(0, 10, titel, ln=True)

    pdf.ln(5)
    pdf.set_font("DejaVu", "", 8)

    # Kolomnamen
    for i, col in enumerate(df.columns):
        if isinstance(col, (pd.Timestamp, datetime)):
                waarde = col.strftime("%Y-%m-%d")
        else:
                waarde = str(col)
        if i == 0:
            pdf.cell(120, 8, waarde, border=1)
        else:
            pdf.cell(25, 8, waarde, border=1)
        
    pdf.ln()

    # Gegevens
    for _, row in df.iterrows():
        for i, item in enumerate(row):
            # Probeer item als datum te herkennen en te formatteren
          
            waarde = str(item).replace(" 00:00:00", "")
           
            if i == 0:
                pdf.cell(120, 8, waarde, border=1)
            else:
                pdf.cell(25, 8, waarde, border=1)
        pdf.ln()

    if dfproblem.empty:
        pdf.cell(0, 10, "Geen geregistreerde opmerkingen voor deze maand.", ln=True)
    else:
        pdf.cell(0, 10, "", ln=True)
        pdf.cell(0, 10, "Overzicht opmerkingen voor deze maand:", ln=True)
        for x, row in dfproblem.iterrows():
            datum = row['Datum'].strftime('%Y-%m-%d')
            opmerking = str(row['Opmerking']).replace(" 00:00:00", "")
            
            pdf.cell(30, 8, datum, border=1)
            pdf.cell(200, 8, opmerking, border=1)
                    #pdf.multi_cell(0, 8, f"🗓️ {datum} – {opmerking}", border=0)
            pdf.ln()
  

    # 💾 Wegschrijven naar geheugen
    output = BytesIO()
    pdf_bytes = pdf.output(dest="S").encode("latin1")
   
    output.write(pdf_bytes)
    output.seek(0)
    return output

##body

st.set_page_config(page_title="Overzicht geregistreerde taken", layout="wide")

st.title("📋 Overzicht geregistreerde taken")

# Upload bestand
taken_file = st.file_uploader("**Upload hier je excel met alle taken**", type=["xlsx"])
problemen_file = st.file_uploader("**Upload hier je excel met problemen**", type=["xlsx"])

if taken_file:
    df = pd.read_excel(taken_file)
    df['Datum'] = pd.to_datetime(df['Datum'])
    df['JaarMaand'] = df['Datum'].dt.strftime("%Y-%m")

    
    # Kies klant en maand
    klant_opties = df['Klant'].unique()
    maand_opties = sorted(df['JaarMaand'].unique(), reverse=True)

    klant = st.selectbox("Kies klant", klant_opties)
    maand = st.selectbox("Kies maand", maand_opties)

    # Filter
    df_filtered = df[(df['Klant'] == klant) & (df['JaarMaand'] == maand)]
    if problemen_file: 
        dfprob = pd.read_excel(problemen_file)
        dfprob['Datum'] = pd.to_datetime(dfprob['Datum'])
        dfprob['JaarMaand'] = dfprob['Datum'].dt.strftime("%Y-%m")
        df_problemen = dfprob[(dfprob['Titel'] == klant) & (dfprob['JaarMaand'] == maand)]

    df_done = df_filtered[df_filtered['Uitgevoerd'] == 1]
    extra_kolom = "Frequentie"
    overzicht = pd.pivot_table(
    df_filtered,
    index='Taak',
    columns='Datum',
    values='Uitgevoerd',
    aggfunc='max',  # toont True als minstens één keer uitgevoerd
    fill_value=False
    )
    
    # Zet True in ✓ en False in ""
    overzicht = overzicht.applymap(lambda x: "✓" if x else "")

    # Haal unieke taak-info op
    taak_info = df_filtered[['Taak', extra_kolom]].drop_duplicates().set_index('Taak')

    # Voeg extra info toe aan matrix
    overzicht = taak_info.join(overzicht)

    # Reset index zodat 'Taak' ook een kolom wordt in de downloadbare Excel
    overzicht = overzicht.reset_index()

    # Sorteervolgorde voor Frequentie
    frequentie_volgorde = ["Dagelijks", "Meerdere keren per week", "Wekelijks", "Tweewekelijks", "Maandelijks", "Jaarlijks", "Op afroep"]

    # Voeg hulpkolom toe met sorteerwaarde (0 = hoogste frequentie)
    overzicht["Frequentie_waarde"] = overzicht["Frequentie"].apply(
    lambda x: frequentie_volgorde.index(x) if x in frequentie_volgorde else len(frequentie_volgorde)
    )

    # Sorteer op Frequentie en verwijder de hulpkolom
    overzicht = overzicht.sort_values(by="Frequentie_waarde").drop(columns="Frequentie_waarde")


    # Resultaat tonen
    st.markdown(f"### 📊 Overzicht voor **{klant}** – **{maand}**")
    st.dataframe(overzicht)
   
    if problemen_file: 
        st.markdown(f"### 📊 Overzicht problemen")
        st.dataframe(df_problemen)


    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from io import BytesIO
    import calendar

    # 🗓️ Titel opbouwen
    titel = f"Overzicht voor {klant} – {maand}"

    # 📄 Maak nieuw Excel-bestand
    wb = Workbook()
    ws = wb.active

    # 📌 Titel toevoegen (eerste rij)
    ws.append([titel])
    ws.append([])  # Lege rij

    # 📊 DataFrame toevoegen vanaf rij 3
    for r in dataframe_to_rows(overzicht, index=False, header=True):
        ws.append(r)

    from openpyxl.styles import numbers

    # Pas datumformaat toe op datum-kolommen
    for col in ws.iter_cols(min_row=3, max_row=ws.max_row):  # vanaf rij 3 (data)
        if isinstance(col[0].value, (pd.Timestamp, pd._libs.tslibs.timestamps.Timestamp)):
            for cell in col:
                cell.number_format = numbers.FORMAT_DATE_DDMMYY

    for column_cells in ws.columns:
        lengte = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        kolom_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[kolom_letter].width = lengte + 2  
    
    # 💾 Export naar geheugen
    output = BytesIO()
    wb.save(output)

    # 📥 Downloadknop
    st.download_button("📥 Download als Excel", output.getvalue(), file_name="taken_overzicht.xlsx")

   
    # PDF-genereerknop
    pdf_bytes = generate_pdf(titel, overzicht, df_problemen)
    st.download_button("📄 Download als PDF", data=pdf_bytes.getvalue(), file_name="taken_overzicht.pdf", mime="application/pdf")




